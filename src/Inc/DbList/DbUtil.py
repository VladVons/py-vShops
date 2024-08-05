# Created: 2024.08.03
# Author: Vladimir Vons <VladVons@gmail.com>
# License: GNU, see LICENSE for more details


from io import BytesIO
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
#
from Inc.Util.Obj import DeepGetByList
from .DbList import TDbList


def DblToXlsx(aDbl: list[TDbList]) -> bytes:
    WB = Workbook()
    WB.remove(WB.active)
    for Idx, xDbl in enumerate(aDbl):
        Title = DeepGetByList(xDbl.Tag, ['name'], f'Sheet{Idx+1}')
        WS = WB.create_sheet(title=Title)

        RowNo = 1
        Fields = xDbl.GetFields()
        Rec = xDbl.RecGo(0)
        for Idx, xField in enumerate(Fields):
            Cell = WS.cell(RowNo, Idx + 1)
            Cell.value = xField
            Cell.font = Font(bold = True)

            CD = WS.column_dimensions[get_column_letter(Idx + 1)]

            FieldVal = Rec.GetField(xField)
            if (isinstance(FieldVal, int)):
                CD.number_format = '#0'
            elif (isinstance(FieldVal, float)):
                CD.number_format = '#,##0.00'

            FieldFmt = DeepGetByList(xDbl.Tag, ['fields', xField])
            if (FieldFmt):
                Val = FieldFmt.get('name')
                if (Val):
                    Cell.value = Val

                Val = FieldFmt.get('format')
                if (Val):
                    CD.number_format = Val

                Val = FieldFmt.get('width')
                if (Val):
                    CD.width = Val

        RowNo += 1
        WS.freeze_panes = WS.cell(RowNo, 1)
        for Rec in xDbl:
            for Idx, Field in enumerate(Fields):
                Cell = WS.cell(RowNo + xDbl.RecNo, Idx + 1).value = Rec.GetField(Field)

    Stream = BytesIO()
    WB.save(Stream)
    Res = Stream.getvalue()
    Stream.close()
    return Res
