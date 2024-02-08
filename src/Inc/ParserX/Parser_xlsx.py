# Created: 2022.09.24
# Author: Vladimir Vons <VladVons@gmail.com>
# License: GNU, see LICENSE for more details


from openpyxl import load_workbook
#
from .Common import TEngine


class TParser_xlsx(TEngine):
    def _InitEngine(self, aFile: str):
        return load_workbook(aFile, read_only = not True, data_only = True) # read_only is False for row_dimensions

    def _GetHeader(self, aRow) -> dict:
        Res = {}
        for x in aRow:
            if (x.value):
                assert (x.value not in Res), f'Field {x.value} not uniq'
                Res[x.value] = x.column
        return Res

    async def _Load(self):
        assert(self._Engine), f'InitEngine() not invoked {self.GetFile()}'

        if (self._Sheet == 'default'):
            WSheet = self._Engine.active
        else:
            assert(self._Sheet in self._Engine.sheetnames), f'Sheet not found {self._Sheet}'
            WSheet = self._Engine[self._Sheet]

        Conf = self.GetConfSheet()
        ConfFields = Conf.get('fields')
        ConfSkip = Conf.get('skip', 0)
        ConfHeader = Conf.get('header', -1)

        if (ConfHeader > 0):
            Header = self._GetHeader(WSheet[ConfHeader])
            for Field, Title in ConfFields.items():
                ConfFields[Field] = Header[Title]

        for RowNo, Row in enumerate(WSheet.rows):
            if (RowNo >= ConfSkip) and (not WSheet.row_dimensions[RowNo+1].hidden):
                Data = {'no': RowNo}
                for Field, FieldIdx in ConfFields.items():
                    if (FieldIdx <= len(Row)):
                        Val = Row[FieldIdx - 1].value
                    else:
                        Val = None
                    Data[Field] = Val
                self._Fill(Data)
                await self.Sleep.Update()

    def _Fill(self, aRow: dict):
        raise NotImplementedError()
